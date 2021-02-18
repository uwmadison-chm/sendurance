import sys
import argparse
import coloredlogs
import logging

from datetime import datetime, timedelta
import dateutil.parser

import csv
import openpyxl

import json
from fitbit_api import FitbitApi
from download_wrapper import DownloadWrapper

# An example of reading from a complex Excel worksheet to auto-download fitbit data

CHUNK_WIDTH = 5


parser = argparse.ArgumentParser()
parser.add_argument('-v', '--verbose', action='count')
parser.add_argument('-o', '--output', default='./raw', help='Path to store output in')
parser.add_argument('input', help='Input CSV listing participant IDs and fitbit account emails')
args = parser.parse_args()

if args.verbose:
    if args.verbose > 1:
        coloredlogs.install(level='DEBUG')
    elif args.verbose > 0:
        coloredlogs.install(level='INFO')
else:
    coloredlogs.install(level='WARN')

try:
    with open("client.json") as json_file:  
        client = json.load(json_file)
except FileNotFoundError as e:
    sys.exit("You need to make sure there is a client.json file in the current directory with an 'id' and 'secret' keys'. If all your fitbits use the same password, also add a 'password' key.")


def read_data(file_path, sheet_name="Distribution"):
    workbook = openpyxl.load_workbook(filename=file_path)
    sheet = workbook[sheet_name]
    row_list = []

    # The data is crazily laid out in 5 column chunks, so we need to restructure it.
    chunks = int(sheet.max_column / CHUNK_WIDTH)

    for chunk in range(0, chunks):
        for row_index in range(3, sheet.max_row):
            start_col = chunk * CHUNK_WIDTH + 1
            fitbit_id = sheet.cell(row=row_index, column=start_col).value
            if not fitbit_id:
                continue

            ident_value = sheet.cell(row=row_index, column=start_col + 1).value
            if not ident_value:
                continue

            start_date = sheet.cell(row=row_index, column=start_col + 2).value
            end_date = sheet.cell(row=row_index, column=start_col + 3).value

            if not start_date:
                logging.warning(f"Skipping {ident_value}, missing start date")
                continue

            if not end_date:
                logging.warning(f"Skipping {ident_value}, missing end date")
                continue

            try:
                ident = int(ident_value)
                d = {
                    'fitbit_id': fitbit_id,
                    'id': ident,
                    'start_date': start_date - timedelta(days=1),
                    'end_date': end_date + timedelta(days=1),
                }
                logging.debug(d)
                row_list.append(d)
            except:
                logging.warning(f"Failure on identifier {ident_value}")

    return sorted(row_list, key=lambda x: x['id'])


r = read_data(args.input)
for row in r:
    fitbit_id = row['fitbit_id']
    ppt = row['id']
    start_date = row['start_date']
    end_date = row['end_date']

    if '*' in fitbit_id:
        # Special password, check for it in secret client hash
        fitbit_id = fitbit_id.strip('*')
        if fitbit_id in client['password-overrides']:
            password = client['password-overrides'][fitbit_id]
        else:
            logging.warning(f"Skipping row with * in fitbit id {fitbit_id}, ppt {ppt}")
            continue
    else:
        password = client['password']

    email = f'afc.fitbit+{fitbit_id}@gmail.com'

    if ppt >= 3000:
        logging.warning(f"Skipping pilot ppt {ppt} with fitbit id {fitbit_id}")
        continue

    logging.info(f"Initializing connection for {ppt}, fitbit account {email} with internal id {fitbit_id}, between {start_date} and {end_date}")
    fitbit = FitbitApi(email,
            password, client['id'], client['secret'],
            debug=args.verbose > 1)
    downloader = DownloadWrapper(ppt=ppt, fitbit=fitbit, start=start_date,
                                 end=end_date, output=args.output)
    downloader.save_hrv()
    downloader.save_sleep_summary()
    downloader.save_sleep_details()
    downloader.save_steps()
